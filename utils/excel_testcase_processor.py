#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'YinJia'

import logging

import openpyxl
from openpyxl.styles import Font, Alignment


class ExcelTestCaseProcessor:
    """文件写入数据"""
    row_number = "row_number"
    sheet_name = "sheet_name"

    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath)
        # if not os.path.exists(self.filename):
        #     # 文件不存在，则拷贝模板文件至指定报告目录下
        #     shutil.copyfile(setting.SOURCE_FILE, setting.TARGET_FILE)

        # self.workbook = self.workbook.get_sheet_by_name(sheet_name)
        # self.workbook = self.wb.active()

    def read_data(self):
        all_data = []
        for sheet in self.workbook.sheetnames:
            work_sheet = self.workbook[sheet]
            data = []
            row_no = 1
            row0 = None
            for row in work_sheet.iter_rows(values_only=True):
                if row_no == 1:
                    row0 = row
                elif row and row[0] is not None:  # 确保行不为空
                    # row_cell = work_sheet[row_no] # 一行单元格对象的元组
                    # row_dict = dict(zip(row_cell, row))
                    row_dict = dict(zip(row0, row))
                    # 在字典中添加行号，行号从1开始
                    row_dict[self.row_number] = row_no
                    row_dict[self.sheet_name] = sheet
                    logging.debug(row_dict)
                    data.append(row_dict)
                row_no += 1
            all_data.extend(data)

        return all_data

    # 定义颜色常量
    GREEN = '008000'
    RED = 'FF0000'
    DARKYELLOW = 'FFCC00'
    font_green = Font(name='宋体', color=GREEN, bold=True)
    font_red = Font(name='宋体', color=RED, bold=True)
    font_yellow = Font(name='宋体', color=DARKYELLOW, bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')

    def write_data(self, row_data, value="PASS"):
        """
        写入测试结果
        :param row_data:一行数据
        :param value: 测试结果值
        :return: 无
        """
        # 获取所在行和列的单元格
        worksheet = self.workbook[row_data[self.sheet_name]]
        row = row_data[self.row_number]

        column = 9
        if "result" in row_data:
            # 获取键的插入顺序位置
            column = list(row_data.keys()).index("result") + 1

        cell = worksheet.cell(row=row, column=column)
        if value == "PASS":
            # 在设置值之后设置字体
            cell.value = value
            cell.font = self.font_green
        elif value == "FAIL":
            cell.value = value
            cell.font = self.font_red  # 在设置值之后设置字体
        worksheet.cell(row=row, column=column).alignment = self.alignment_center

        # 写入作者
        worksheet.cell(row=row, column=column + 1, value="Jason")
        # 设置字体和对其方式
        worksheet.cell(row=row, column=column + 1).font = self.font_yellow
        worksheet.cell(row=row, column=column + 1).alignment = self.alignment_center

        # 保存工作簿
        self.workbook.save(filename=self.filepath)
